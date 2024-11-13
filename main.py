import speech_recognition as sr
import pyttsx3
import re
import spacy
import streamlit as st
import pandas as pd
import openpyxl
import time

# Load spaCy's English language model
nlp = spacy.load("en_core_web_sm")

# Initialize the recognizer and the text-to-speech engine
r = sr.Recognizer()
engine = pyttsx3.init()

# Set pause threshold to 2 seconds to wait for a 2-second pause before stopping
r.pause_threshold = 2.0

# Function to extract information from the recognized text using NLP
def extract_information(text):
    doc = nlp(text)

    information = {
        "Name of the patient": None,
        "Age": None,
        "Gender": None,
        "Phone Number": None,
        "Date of Birth": None,
        "Address": None
    }

    # Regex patterns to match specific fields
    patterns = {
        "Name of the patient": [r"name(?: of the)? patient\s*(?:is)?\s*([a-zA-Z\s]+)(?=\s*(age|gender|phone number|address|$))"],
        "Age": [r"age\s*(?:is)?\s*(\d+)", r"\b(\d+)\s+years?\s+old\b"],
        "Gender": [r"gender\s*(?:is)?\s*(\w+)", r"patient\s+is\s+(?:a\s+)?(\w+)\b"],
        "Phone Number": [r"phone number\s*(?:is)?\s*([\d\s]+)", r"contact\s+number\s*(?:is)?\s*([\d\s]+)"],
        "Date of Birth": [r"(?:date of birth|birth date)\s*(?:is)?\s*(.+?)(?=\s*(address|$))", r"born\s+on\s*(.+?)(?=\s*(address|$))"],
        "Address": [r"address\s*(?:is)?\s*(.+)", r"lives\s+at\s*(.+)"]
    }

    # Match regex patterns
    for field, patterns_list in patterns.items():
        if information[field] is None:
            for pattern in patterns_list:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    information[field] = match.group(1).strip()
                    break

    return information

# Function to listen for speech for exactly 15 seconds
def start_listening():
    st.write("Listening for speech...")

    with sr.Microphone() as source:
        r.adjust_for_ambient_noise(source, duration=1)

        start_time = time.time()  # Record the start time
        audio_data = []

        while time.time() - start_time < 15:  # Listen for 15 seconds
            try:
                # Listen for audio and append it to the audio_data list
                audio_data.append(r.listen(source, timeout=0.1))
            except sr.WaitTimeoutError:
                continue  # If no speech detected, just continue waiting

        # Combine all the audio fragments captured during the 15 seconds
        full_audio_data = sr.AudioData(b''.join([a.frame_data for a in audio_data]), audio_data[0].sample_rate, audio_data[0].sample_width)

        # Store the full audio data in session state
        st.session_state['audio_data'] = full_audio_data
        st.success("Audio captured successfully! Now processing...")

# Function to process speech after 15 seconds of listening
def process_audio_after_15_sec():
    if 'audio_data' not in st.session_state:
        st.warning("No audio captured. Please press 'Start Speech' first.")
        return None

    audio_data = st.session_state['audio_data']

    try:
        st.write("Processing speech...")

        # Convert audio to text
        text = r.recognize_google(audio_data)
        st.write(f"Recognized Speech: {text}")

        if not text.strip():
            st.warning("No valid speech detected.")
            return None

        # Extract information from recognized speech
        extracted_info = extract_information(text)

        # Create a DataFrame for the extracted information
        df = pd.DataFrame([extracted_info])
        
        st.write("Extracted Information:")
        st.dataframe(df)

        return df

    except sr.UnknownValueError:
        st.error("Could not understand the audio")
    except sr.RequestError as e:
        st.error(f"Could not request results from the speech recognition service; {e}")
    except Exception as e:
        st.error(f"An unexpected error occurred: {e}")

# Function to save patient data to Excel
def save_patient_data_excel(df):
    try:
        # Specify the directory where the Excel file is located
        excel_file_path = 'patient_data.xlsx'

        # Load the workbook
        wb = openpyxl.load_workbook(excel_file_path)
        sheet = wb.active  # Get the active sheet

        # Generate Unique ID
        last_row = sheet.max_row
        if last_row == 1:  # If the sheet is empty (only headers)
            new_id = 1
        else:
            # Get the last ID, handle None values
            last_id = sheet.cell(row=last_row, column=1).value
            new_id = 1 if last_id is None else int(last_id) + 1

        # Add the new ID and Prescription No. to the DataFrame
        df['ID'] = new_id

        # Reorder columns to match the Excel file
        df = df[['ID', 'Name of the patient', 'Age', 'Gender', 'Phone Number',
                 'Date of Birth', 'Address']]

        # Append data to the sheet
        for index, row in df.iterrows():
            sheet.append(list(row.values))

        wb.save(excel_file_path)
        st.success(f"Patient data saved to Excel successfully at: {excel_file_path}")

    except Exception as e:
        st.error(f"Error saving patient data to Excel: {e}")

# Function to load patient data from Excel
def load_patient_data():
    try:
        excel_file_path = 'patient_data.xlsx'
        wb = openpyxl.load_workbook(excel_file_path)
        sheet = wb.active

        data = sheet.values
        columns = next(data)[0:]  # Get the first row as column headers
        df = pd.DataFrame(data, columns=columns)
        return df

    except Exception as e:
        st.error(f"Error loading patient data from Excel: {e}")
        return None

# Function to update patient data in Excel
def update_patient_data_excel(df, patient_id, updated_data):
    try:
        excel_file_path = 'patient_data.xlsx'
        wb = openpyxl.load_workbook(excel_file_path)
        sheet = wb.active

        # Find the row to update
        row_to_update = None
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from row 2 to skip headers
            if row[0] == patient_id:
                row_to_update = row[0]
                break

        if row_to_update is None:
            st.error(f"Patient with ID {patient_id} not found.")
            return

        # Update the data in the row
        for col_num, field in enumerate(updated_data.keys(), 1):
            sheet.cell(row=row_to_update, column=col_num).value = updated_data[
                field]

        wb.save(excel_file_path)
        st.success("Patient data updated successfully!")

    except Exception as e:
        st.error(f"Error updating patient data to Excel: {e}")





import re

def extract_diagnosis_info(text):
    # Initialize dictionary for symptoms, treatment, and medicine
    diagnosis_info = {
        "Symptoms": None,
        "Treatment": None,
        "Medicine": None
    }

    # Define regex patterns to capture complex sentences
    patterns = {
        "Symptoms": r"(symptoms?|fever|pain|cough|headache|mild\s+fever|nausea|vomiting)[^\w]*(.*?)(?=\s*(treatment|medicine|$))",
        "Treatment": r"(treatment|suggestion|care|recommendation)\s*[:\-\-]?\s*(.*?)(?=\s*(medicine|$))",
        "Medicine": r"(medicine|medication)\s*[:\-\-]?\s*(.*?)(?=\s*(symptoms|treatment|$))"
    }

    # Iterate over each field and match using the regex
    for field, pattern in patterns.items():
        match = re.search(pattern, text.lower())  # Convert to lower case for case-insensitive matching
        if match:
            diagnosis_info[field] = match.group(2).strip()  # Extract the content for the field
            # Remove matched content from text to avoid overlapping with other fields
            text = text.replace(match.group(0), "")  # Remove matched text to avoid overlap

    # Ensure no overlap occurs, and each field is uniquely populated
    if diagnosis_info["Symptoms"] is None:
        symptoms_match = re.match(r"([a-zA-Z\s,]+?)(?=\s*(treatment|medicine|$))", text.strip())
        if symptoms_match:
            diagnosis_info["Symptoms"] = symptoms_match.group(1).strip()

    if diagnosis_info["Treatment"] is None and diagnosis_info["Symptoms"]:
        treatment_match = re.match(r"(.*?)(?=\s*medicine|\s*$)", text.strip())
        if treatment_match:
            diagnosis_info["Treatment"] = treatment_match.group(1).strip()

    if diagnosis_info["Medicine"] is None:
        medicine_match = re.match(r"(.*)", text.strip())
        if medicine_match:
            diagnosis_info["Medicine"] = medicine_match.group(1).strip()

    return diagnosis_info



def capture_diagnosis():
    st.write("Listening for doctor's input...")

    with sr.Microphone() as source:
        r.adjust_for_ambient_noise(source, duration=1)
        audio_data = r.listen(source, timeout=15)
    
    try:
        text = r.recognize_google(audio_data)
        st.write(f"Recognized Speech: {text}")  # Debug: Show recognized speech
        
        # If no text is recognized, show a message
        if not text.strip():
            st.warning("No speech detected.")
            return None
        
        # Extract diagnosis info from recognized speech
        diagnosis_info = extract_diagnosis_info(text)
        
        # Show extracted diagnosis info for debugging
        st.write(f"Extracted Diagnosis Info: {diagnosis_info}")
        
        # Check if the extracted information is valid
        if any(value is not None for value in diagnosis_info.values()):
            diagnosis_df = pd.DataFrame([diagnosis_info])
            st.write("Diagnosis Information Captured:")
            st.dataframe(diagnosis_df)  # Display the extracted DataFrame
            return diagnosis_df
        else:
            st.warning("No valid diagnosis information extracted.")
            return None

    except sr.UnknownValueError:
        st.error("Could not understand the audio.")
    except sr.RequestError as e:
        st.error(f"Could not request results from the speech recognition service; {e}")
    except Exception as e:
        st.error(f"An unexpected error occurred: {e}")


import openpyxl
import pandas as pd
import streamlit as st

def save_diagnosis_data(patient_id, diagnosis_df):
    try:
        excel_file_path = 'patient_data.xlsx'
        wb = openpyxl.load_workbook(excel_file_path)
        sheet = wb.active

        # Find the row to update based on the patient ID
        row_to_update = None
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):  # Start from row 2 to skip headers
            if row[0].value == patient_id:  # Assuming Patient ID is in the first column (column 0 in openpyxl indexing)
                row_to_update = row[0].row  # Get the row number
                break

        if row_to_update is None:
            st.error(f"Patient with ID {patient_id} not found.")
            return

        # Update diagnosis columns (assuming columns for Symptoms, Treatment, and Medicine are columns 8, 9, and 10)
        if not diagnosis_df.empty:
            sheet.cell(row=row_to_update, column=8).value = diagnosis_df['Symptoms'].iloc[0]  # Column 8 for Symptoms
            sheet.cell(row=row_to_update, column=9).value = diagnosis_df['Treatment'].iloc[0]  # Column 9 for Treatment
            sheet.cell(row=row_to_update, column=10).value = diagnosis_df['Medicine'].iloc[0]  # Column 10 for Medicine

        wb.save(excel_file_path)
        st.success("Diagnosis data saved successfully!")

    except Exception as e:
        st.error(f"Error saving diagnosis data to Excel: {e}")


# Function to switch between portals
def show_sidebar():
    st.sidebar.title("Portal Navigation")
    return st.sidebar.radio("Select Portal", ["Admin Portal", "Doctor's Portal", "Patient Portal", "Chatbot"])


def admin_portal():
    # Initialize session state for audio data if not already present
    if 'audio_data' not in st.session_state:
        st.session_state['audio_data'] = None

    # Option to add new patient or edit existing patient
    operation = st.radio("Select Operation:",
                         ("Add New Patient", "Edit Existing Patient"))

    if operation == "Add New Patient":
        # Buttons to start and stop speech 
        if st.button('Start Speech'):
            start_listening()

        if st.button('Stop Speech and Process'):
            df = process_audio_after_15_sec()
            if df is not None:
                save_patient_data_excel(df)

    elif operation == "Edit Existing Patient":
        df = load_patient_data()
        if df is not None:
            patient_ids = df['ID'].tolist()
            selected_patient_id = st.selectbox("Select Patient ID:", patient_ids)

            # Get the data for the selected patient
            patient_data = df[df['ID'] == selected_patient_id].iloc[0].to_dict()

            # Display current data in a form
            updated_data = {}
            for field, value in patient_data.items():
                if field != 'ID' and field != 'Prescription No.':  # Exclude ID and Prescription No. from editing
                    updated_data[field] = st.text_input(field, value)

            if st.button("Update Patient Data"):
                update_patient_data_excel(df, selected_patient_id, updated_data)



# Doctor's Portal: Fetch patient by ID and enter symptoms, treatment, medicine
def doctors_portal():
    st.title("Doctor's Portal")
    
    # Load patient data
    df = load_patient_data()
    
    if df is not None:
        patient_ids = df['ID'].tolist()
        selected_patient_id = st.selectbox("Select Patient ID:", patient_ids)
        
        if selected_patient_id:
            # Display selected patient details
            patient_data = df[df['ID'] == selected_patient_id].iloc[0].to_dict()
            st.write("Patient Details:")
            for field, value in patient_data.items():
                st.write(f"{field}: {value}")

            # Start diagnosis entry
            if st.button('Start Diagnosis Speech'):
                # Capture diagnosis information from speech
                diagnosis_df = capture_diagnosis()
            
            
                
                # Display the captured diagnosis info in Streamlit
                if diagnosis_df is not None and not diagnosis_df.empty:
                    st.write("Diagnosis Information Captured:")
                    st.dataframe(diagnosis_df)
                    
                    # Save diagnosis data to the Excel file
                    save_diagnosis_data(selected_patient_id, diagnosis_df)
















# Patient Portal: View own details (for demo)
def patient_portal():
    st.header("Patient Portal")
    df = load_patient_data()
    if df is not None:
        patient_ids = df['ID'].tolist()
        selected_patient_id = st.selectbox("Select Your Patient ID:", patient_ids)
        patient_data = df[df['ID'] == selected_patient_id].iloc[0].to_dict()
        st.write("Your Details:", patient_data)


from chat_page import display_chat


# Streamlit App UI
def app():
    st.title("Doctor's Prescription Form")
    portal_choice = show_sidebar()
    if portal_choice == "Admin Portal":
        admin_portal()
    elif portal_choice == "Doctor's Portal":
        doctors_portal()
    elif portal_choice == "Patient Portal":
        patient_portal()
    elif portal_choice == "Chatbot":
        display_chat()
    
# Run the Streamlit app
if __name__ == "__main__":
    app()